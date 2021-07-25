import openpyxl

if __name__ == '__main__':
    # 打开一个wk文件
    exampleXls = openpyxl.load_workbook('example.xlsx')
    print(type(exampleXls))
    print('-' * 20 + '\n')


    # 获取它的sheet页
    sheets = exampleXls.worksheets
    print(sheets)
    print('-' * 20 + '\n')


    # 选择其中第一个sheet
    wkSheet = sheets[0]
    print(type(wkSheet))
    print('sheet名：' + wkSheet.title)
    print('行数：' + str(wkSheet.max_row))
    print('列数：' + str(wkSheet.max_column))
    print('-' * 20 + '\n')


    # 从sheet中获取一个cell
    cellA1 = wkSheet['A1']
    print('单元格值: ' + str(cellA1.value))
    # 另一种方式获取cell
    cellA1 = wkSheet.cell(1, 1)
    print('单元格值: ' + str(cellA1.value))


    # 遍历第一行数据
    for columnIndex in range(1, wkSheet.max_column + 1):
        cell = wkSheet.cell(1, columnIndex)
        print(cell.coordinate + ' : ' + str(cell.value))
    print('-' * 20 + '\n')

    # 获取某一区域
    ranges = wkSheet['A1:C3']
    for row in ranges:
        for cell in row:
            print(cell.coordinate, cell.value)
    print('-' * 20 + '\n')

