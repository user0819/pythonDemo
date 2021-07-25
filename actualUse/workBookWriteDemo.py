import openpyxl
import openpyxl.styles

if __name__ == '__main__':
    # 新建一个wk文件
    exampleXls = openpyxl.Workbook()
    print(type(exampleXls))
    print('-' * 20 + '\n')


    # 移除默认自带的sheet
    exampleXls.remove(exampleXls.worksheets[0])
    print('删除sheet')
    print('-' * 20 + '\n')


    # 创建一个新的sheet
    newSheet = exampleXls.create_sheet('mySheet')
    sheets = exampleXls.worksheets
    print(sheets)
    print('-' * 20 + '\n')

    # 为新sheet的cell赋值
    cell = newSheet['A1']
    cell.value = 'newValue'


    # 设置行高、列宽
    newSheet.row_dimensions[1].height = 70
    newSheet.column_dimensions['A'].width = 20


    # 合并和拆分单元格
    newSheet['A3'] = '合并单元格1'
    newSheet['C4'] = '合并单元格2'
    newSheet.merge_cells('A3:C4')

    #newSheet.unmerge_cells('C5:C6')


    # 保存
    exampleXls.save('writeDemo.xlsx')




